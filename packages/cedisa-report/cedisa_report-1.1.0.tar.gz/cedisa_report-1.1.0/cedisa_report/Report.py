import datetime
from openpyxl import drawing, utils
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.styles import Border, Side, PatternFill

from src.ReportAdapter import ReportAdapter

class Report(ReportAdapter):

    def __init__(self, ws: Worksheet) -> None:
        self.title_text_color = '00123c'
        self.orange_text_color = 'dc3002'  #Talvez sirva para denotar números negativos
        self.header_color = 'c0c0c0'
        self.text_color = '333333'

        self.thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        self.medium_border = Border(left=Side(style='medium'), 
                            right=Side(style='medium'), 
                            top=Side(style='medium'), 
                            bottom=Side(style='medium'))

        self.header_bg = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type="solid")
        self.ws = ws

    def make_title(self, title: str, merge_range: str) -> None:
        self.ws.append([title])
        self.ws.merge_cells(merge_range)
        self.ws.row_dimensions[1].height = 60
        title_font_style = Font('Arial', 18, True, color=self.title_text_color, vertAlign='superscript')
        img = drawing.image.Image('./img/logo-esticada.png')
        img.anchor = 'A1'
        img.height = 70
        img.width = 180
        self.ws.add_image(img)
        titulo = self.ws.cell(1, 1)
        titulo.alignment = Alignment(vertical='center', horizontal='center')
        titulo.font = title_font_style
        for i in range(11):
            if i != 0:
                self.ws.cell(1, i).border = self.thin_border        
    
    def make_report_creation_date(self) -> None:
        day = datetime.datetime.now().day if datetime.datetime.now().day >= 10 else '0' + str(datetime.datetime.now().day)
        month = datetime.datetime.now().month if datetime.datetime.now().month >= 10 else '0' + str(datetime.datetime.now().month)
        yerar = datetime.datetime.now().year
        text = f'Gerado em: {day}/{month}/{yerar}'
        font_style = Font('Arial', 14, True, color=self.text_color, vertAlign='superscript')
        self.ws.cell(1, 11).font = font_style
        self.ws.cell(1, 11).border = self.thin_border
        self.ws.column_dimensions[utils.get_column_letter(11)].width = 20
        self.ws['K1'] = text
    
    def make_simple_header(self, title: str, merge_range: str, line_number: int, column_number: int, width: int = 30) -> None:
        self.ws.merge_cells(merge_range)
        font_style = Font('Arial', 14, True, color=self.text_color, vertAlign='superscript', bold=True)
        cell = self.ws.cell(line_number, column_number)
        # cell.alignment.wrap_text = True
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell.fill = self.header_bg
        cell.font = font_style
        cell.value = title

        if width != None:
            self.ws.column_dimensions[utils.get_column_letter(column_number)].width = width

        for i in range(line_number + 2):
            if i >= line_number:
                self.ws.cell(i, column_number).border = self.thin_border
    
    def make_text_data(self, line_number: int, column_number: int, data) -> None:
        font_style = Font('Arial', 14, True, color=self.text_color, vertAlign='superscript', bold=True)
        cell = self.ws.cell(line_number, column_number)
        cell.font = font_style
        if data != None: cell.value = data
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def make_number_data(self, line_number: int, column_number: int, data) -> None:
        font_style = Font('Arial', 14, True, color=self.text_color, vertAlign='superscript', bold=True)
        cell = self.ws.cell(line_number, column_number)
        cell.font = font_style
        if data != None: cell.value = data
        cell.border = self.thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')