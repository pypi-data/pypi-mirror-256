# -*- coding : utf-8-*-
import os
import pandas as pd
import polars as pl
import xlwings as xw
import datetime
from xlsx2csv import Xlsx2csv
from chardet.universaldetector import UniversalDetector
"""
v0.0.1第一个版本
v0.0.2
修改了polar引擎，在read_csv_options中增加了{"infer_schema_length":1000,"ignore_errors":False})两个配置，性能进一步提升，特殊情况下兼容性降低，
可以采用pandas引擎弥补（之前需要对文件读两次完成类型推断，一次类型推断，一次读文件。当前只对前1000行读进行内容推断）。
v0.0.3
1、更新openpyxl引擎写的load_excel函数，比pandas引擎快20%，兼容性基本一致。该函数可以支持部分行读取。
2、polars引擎增加了read_csv_options选项，目前默认关闭。打开后使用者可以自己设置各类参数，主要是{"infer_schema_length":1000,"ignore_errors":False} 该参数影响类型推断的精度和速度。
3、polars引擎读取excel表后默认删除全部值为空的行和列。
v0.0.4
1、xlwings引擎增加类型转换
2、load函数增加错误提示
v0.0.5
1、增加读取报错时，自动转为pandas 读取,并提供提示。
2、增加追加写功能，分别通过pandas和xlwings实现，其中小数据集通过pandas较为快速方便，对于大数据集或者需要指定写入文件单元格的情况使用xlwings库。
注意：
2.1、追加写时文档必须关闭，否则使用pandas引擎会报错，xlwings不会报错但是无法追加写入
2.2、需要指定写入文件单元格的情况引擎必须使用xlwings库，既engine == "xlwings"
3、增加在文件后缀上自动添加写入时间的功能，时间格式为-年月日_时分
4、增加查看功能，可以在excel中打开DataFrame和文件路径进行查看，方便在jupyter等交互环境中使用。


"""

# 将csv转化为utf8编码
def encode_to_utf8(filename, des_encode):
    # 读取文件的编码方式
    with open(filename, 'rb') as f:
        detector = UniversalDetector()
        for line in f.readlines():
            detector.feed(line)
            if detector.done:
                break
        original_encode = detector.result['encoding']
    # 读取文件的内容
    with open(filename, 'rb') as f:
        file_content = f.read()
    # 修改编码
    file_decode = file_content.decode(original_encode, 'ignore')
    file_encode = file_decode.encode(des_encode)
    with open(filename, 'wb') as f:
        f.write(file_encode)


# 通过xlwings读取表
def xw_open(file_path, sheetname='Sheet1', visible=False):
    # 数据类型可能推断不正确。测试时发现可以正确区分object和数据类型，但是数据类型都推断为float64不能区分int64
    app = xw.App(visible=visible,
                 add_book=False)

    book = app.books.open(file_path)

    sheet = book.sheets[sheetname]
    data = sheet.used_range.options(pd.DataFrame, header=1, index=False, expand='table').value
    data = data.convert_dtypes()
    if visible == False:
        book.close()
        app.quit()

    return data


# 通过xlwings写入表
def xw_write(df, file_path, sheetname='Sheet1', visible=False):
    # 数据类型可能推断不正确。测试时发现可以正确区分object和数据类型，但是数据类型都推断为float64不能区分int64
    app = xw.App(visible=visible,
                 add_book=False)
    wb = app.books.add()

    if sheetname == 'Sheet1':
        sheet = wb.sheets[0]
    else:
        sheet = wb.sheets.add(sheetname, after='Sheet1')

    # 将dataframe写入Sheet
    sheet.range('A1').value = df
    # 保存并关闭Workbook
    wb.save(file_path)
    if visible == False:
        wb.close()
        app.quit()


###通过xlwings追加写入
def xw_write_a(df, file_path, sheetname='Sheet1', cell='A1', visible=False, close=True):
    if not (os.path.exists(file_path)):
        wb = xw.Book()
        wb.save(file_path)

    app = xw.App(visible=visible, add_book=False)
    wb = app.books.open(file_path)

    sheet_names = [sht.name for sht in wb.sheets]
    if sheetname not in sheet_names:
        wb.sheets.add(sheetname, after=sheet_names[-1])

    sheet = wb.sheets[sheetname]
    sheet.range(cell).value = df
    wb.save()
    if close:
        wb.close()
        app.quit()

##通过xlwings查看df
def xw_view(df):
    # 启动Excel程序，不新建工作表薄（否则在创建工作薄时会报错），这时会弹出一个excel
    app= xw.App(visible = True, add_book= False)
    # 新建一个工作簿，默认sheet为Sheet1
    wb = app.books.add()
    #将工作表赋值给sht变量
    sht = wb.sheets('Sheet1')
    sht.range('A1').value =df


###通过pandas追加写入
def pd_write_a(df, file_path, sheetname='Sheet1'):
    with pd.ExcelWriter(file_path, mode="a", engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheetname)


# xlsx转换为csv
def xlsxtocsv(file_path):
    file_path_csv = file_path.replace(".xlsx", ".csv")
    Xlsx2csv(file_path, outputencoding="utf-8").convert(file_path_csv)
    return file_path_csv


# row_count每次读取的行数
def load_stream_row(file_path, row_count, col_name=None):
    name, ext = os.path.splitext(file_path)
    if '.csv' == ext:
        encode_to_utf8(file_path, des_encode="utf-8")
        df_read = pd.read_csv(file_path, usecols=col_name, chunksize=row_count)
    if ".xls" == ext:
        df_read = pd.read_excel(file_path, usecols=col_name)
        # 转化为csv再分块读
        file_path_csv = file_path.replace(".xls", ".csv")
        df_read.to_csv(file_path_csv, index=False, encoding='UTF-8')
        # encode_to_utf8(file_path_csv, des_encode="utf-8")
        df_read = pd.read_csv(file_path_csv, usecols=col_name, chunksize=row_count)
    if ".xlsx" == ext:
        file_path_csv = xlsxtocsv(file_path)
        df_read = pd.read_csv(file_path_csv, usecols=col_name, chunksize=row_count)
    return df_read


#检查路径是否存在，不存在则创建
def check_directory_and_create(file_path):
    directory = os.path.dirname(file_path)
    # 检查目录是否存在
    if not os.path.exists(directory):
        # 创建目录
        os.makedirs(directory)




# 第一行是列名，从第二行开始读内容
def load_excel(file_path, sheetname='Sheet1', start_row=2, end_row=None):
    from openpyxl import load_workbook
    lst = []
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb[sheetname]
    max_row = ws.max_row
    # [*迭代器]方法性能高于list(迭代器)和逐行取值的方法
    row_columns = [*ws.iter_rows(min_row=1, max_row=1, values_only=True)]
    if end_row != None:
        row_data = [*ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True)]
    else:
        row_data = [*ws.iter_rows(min_row=start_row, max_row=max_row, values_only=True)]
    # 将列名和数据合并
    row_columns.extend(row_data)
    df = pd.DataFrame(row_columns)
    return df


########主函数#############################################################################################################

# 自适用后缀、多引擎读取表，默认为polars
def load(file_path, col_name=None, sheetname='Sheet1', engine=None, read_csv_options=None,datatype="pandas"):
    name, ext = os.path.splitext(file_path)
    try:
        if '.csv' == ext:
            if engine == None:
                encode_to_utf8(file_path, des_encode="utf-8")
                df_read = pl.read_csv(file_path, columns=col_name)
                if datatype == "polars":
                    return df_read
                df_read = df_read.to_pandas()
            if engine == "polars":
                encode_to_utf8(file_path, des_encode="utf-8")
                df_read = pl.read_csv(file_path, columns=col_name)
                if datatype == "polars":
                    return df_read
                df_read = df_read.to_pandas()
            if engine == "pandas":
                encode_to_utf8(file_path, des_encode="utf-8")
                df_read = pd.read_csv(file_path, usecols=col_name)
            if engine == "xlwings":
                # xlwings读取csv兼容性和效率都较差调用pandas读取
                encode_to_utf8(file_path, des_encode="utf-8")
                df_read = pd.read_csv(file_path, usecols=col_name)

        if ".xlsx" == ext:
            if engine == None:
                df_read = pd.read_excel(file_path, usecols=col_name, sheet_name=sheetname,engine="calamine")
            if engine == "polars":
                df_read = pl.read_excel(file_path,
                                        read_csv_options=read_csv_options,
                                        sheet_name=sheetname)
                # 删除所有空行
                df_read = df_read.filter(~pl.all(pl.all().is_null()))
                df_read = df_read[[s.name for s in df_read if not (s.null_count() == df_read.height)]]
                df_read = df_read.to_pandas()
            if engine == "pandas":
                df_read = pd.read_excel(file_path, usecols=col_name, sheet_name=sheetname)
            if engine == "xlwings":
                df_read = xw_open(file_path, sheetname=sheetname, visible=False)

        if ".xls" == ext:
            if engine == None:
                df_read = pd.read_excel(file_path, usecols=col_name, sheet_name=sheetname, engine="calamine")
            if engine == "polars":
                # polars不能读xls格式，调用pandas解决
                df_read = pd.read_excel(file_path, usecols=col_name, sheet_name=sheetname)
            if engine == "pandas":
                df_read = pd.read_excel(file_path, usecols=col_name, sheet_name=sheetname)
            if engine == "xlwings":
                df_read = xw_open(file_path, sheetname=sheetname, visible=False)
        if ".pkl" == ext:
            df_read = pd.read_pickle(file_path)

    except Exception as e:
        if '.csv' == ext:
            encode_to_utf8(file_path, des_encode="utf-8")
            df_read = pd.read_csv(file_path, usecols=col_name)

        if ".xlsx" == ext:
            df_read = pd.read_excel(file_path, usecols=col_name, sheet_name=sheetname)

        print(f"读取文件发生错误：{e}")
        print(
            f'已自动切换兼容性更好的pandas引擎。下次读取该文件可以手动选择pandas引擎，语法为load(file_path,engine="pandas")，对于大文件尝试使用engine="xlwings"。')

    return df_read


# 自适应后缀、多引擎写入表,默认为polars，带有追加写入功能
def dump(df_write, file_path, mode=None, sheetname='Sheet1', time=False, engine="polars", cell='A1', visible=False,
         close=True):
    check_directory_and_create(file_path)
    name, ext = os.path.splitext(file_path)
    if time:
        timestamp = datetime.datetime.now().strftime("%y%m%d_%H%M")
        """ 
        strftime="%y%m%d%H%M"  2位数年份 2306010101 Y大写则为4位数年份
        strftime="%m%d%H%M%S"   添加秒的信息 H M S时分秒 不能换为小写字母

        """
        base_path, ext_path = os.path.splitext(file_path)
        file_path_with_time = f"{base_path}%{timestamp}{ext}"
        file_path = file_path_with_time
    try:
        if mode == None:
            if '.csv' == ext:
                if engine == "polars":
                    df_write = pl.from_pandas(df_write)
                    df_write.write_csv(file_path, separator=",")
                if engine == "pandas":
                    df_write.to_csv(file_path, index=False)
                if engine == "xlwings":
                    # xlwings不能写入csv，调用pandas写入
                    df_write.to_csv(file_path, index=False)

            if ".xlsx" == ext:
                if engine == "polars":
                    df_write = pl.from_pandas(df_write)
                    df_write.write_excel(file_path, worksheet=sheetname)
                if engine == "pandas":
                    df_write.to_excel(file_path, index=False, sheet_name=sheetname)
                if engine == "xlwings":
                    xw_write(df_write, file_path, sheetname=sheetname, visible=False)

            if ".xls" == ext:
                if engine == "polars":
                    df_write.to_excel(file_path, index=False, sheet_name=sheetname)
                if engine == "pandas":
                    df_write.to_excel(file_path, index=False, sheet_name=sheetname)
                if engine == "xlwings":
                    xw_write(df_write, file_path, sheetname=sheetname, visible=False)

            if ".pkl" == ext:
                df_write.to_pickle(file_path)
        if mode == "a":
            if '.csv' == ext:
                df_write.to_csv(file_path, index=False, mode='a')

            if ".xlsx" == ext:
                if engine == "polars":
                    pd_write_a(df_write, file_path, sheetname=sheetname)
                if engine == "pandas":
                    pd_write_a(df_write, file_path, sheetname=sheetname)
                if engine == "xlwings":
                    xw_write_a(df_write, file_path, sheetname=sheetname, cell=cell, visible=False, close=True)

            if ".xls" == ext:
                if engine == "polars":
                    pd_write_a(df_write, file_path, sheetname=sheetname,engine="xlsxwriter")
                if engine == "pandas":
                    pd_write_a(df_write, file_path, sheetname=sheetname,engine="xlsxwriter")
                if engine == "xlwings":
                    xw_write_a(df_write, file_path, sheetname=sheetname, cell=cell, visible=False, close=True)
        return file_path
    except Exception as e:
        if '.csv' == ext:
            df_write.to_csv(file_path, index=False)

        if ".xlsx" == ext:
            df_write.to_excel(file_path, index=False, sheet_name=sheetname)

        print(f"写入文件发生错误：{e}")
        print(
            f'已自动切换兼容性更好的pandas引擎。下次写入该文件可以手动选择pandas引擎，语法为dump(file_path,engine="pandas")，对于大文件尝试使用engine="xlwings"。')
        return file_path
    
##通过excel查看数据，输入参数f既可以是文件路径也可以是DataFrame
def view(f):
    if type(f)==str:
        xw_open(f, sheetname='Sheet1', visible=True)
    else:
        xw_view(f)